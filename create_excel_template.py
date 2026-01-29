"""
Excel Template Creator for Book Generation System

This script helps you create a properly formatted Excel file (.xlsx) 
that can be synced to Supabase using sync_from_excel.py

USAGE:
    python create_excel_template.py
    
    This will create 'books.xlsx' with example rows and proper column headers.

WHAT IT DOES:
    Creates an Excel file with the correct column structure for the book generation system.
    You can then edit this file, add your books, and sync it to Supabase.

END GOAL:
    - Create a template Excel file that matches the expected format
    - Make it easy to add multiple books at once
    - Sync all books to Supabase database automatically
    - Then process them through the automated workflow
"""
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logging.basicConfig(level=logging.INFO, format="%(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


def create_excel_template(filename: str = "books.xlsx"):
    """
    Create an Excel template file with proper headers and example rows.
    
    Args:
        filename: Name of the Excel file to create (default: books.xlsx)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Books"

    # Define headers
    headers = [
        "title",
        "notes_on_outline_before",
        "status_outline_notes",
        "chapter_notes_status",
        "final_review_notes_status",
        "notes_on_outline_after",
        "final_review_notes",
        "sheet_row_id",
    ]

    # Header descriptions (for reference)
    descriptions = [
        "Book Title (REQUIRED)",
        "Instructions for outline generation (REQUIRED)",
        "yes/no/no_notes_needed (default: no_notes_needed)",
        "yes/no/no_notes_needed (default: no_notes_needed)",
        "yes/no/no_notes_needed (default: no_notes_needed)",
        "Notes after outline is generated (optional)",
        "Final review notes (optional)",
        "Tracking ID (optional)",
    ]

    # Style header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write descriptions in row 2 (as comments/reference)
    desc_font = Font(italic=True, size=9, color="666666")
    for col_idx, desc in enumerate(descriptions, start=1):
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font = desc_font

    # Example rows
    examples = [
        {
            "title": "AI Agents",
            "notes_on_outline_before": "Make it simple and beginner-friendly. Cover basics of AI, machine learning, and practical applications.",
            "status_outline_notes": "no_notes_needed",
            "chapter_notes_status": "no_notes_needed",
            "final_review_notes_status": "no_notes_needed",
            "notes_on_outline_after": "",
            "final_review_notes": "",
            "sheet_row_id": "",
        },
        {
            "title": "Go Language Programming Guide",
            "notes_on_outline_before": "Cover Python basics, OOP concepts, and advanced topics like decorators and generators.",
            "status_outline_notes": "no_notes_needed",
            "chapter_notes_status": "no_notes_needed",
            "final_review_notes_status": "no_notes_needed",
            "notes_on_outline_after": "",
            "final_review_notes": "",
            "sheet_row_id": "",
        },
        {
            "title": "Data Science with Python",
            "notes_on_outline_before": "Focus on practical data science workflows. Include examples with real datasets.",
            "status_outline_notes": "no_notes_needed",
            "chapter_notes_status": "no_notes_needed",
            "final_review_notes_status": "no_notes_needed",
            "notes_on_outline_after": "",
            "final_review_notes": "",
            "sheet_row_id": "",
        },
    ]

    # Write example rows
    for row_idx, example in enumerate(examples, start=3):
        for col_idx, header in enumerate(headers, start=1):
            value = example.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Style example rows
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Adjust column widths
    column_widths = {
        "A": 25,  # title
        "B": 60,  # notes_on_outline_before
        "C": 20,  # status_outline_notes
        "D": 20,  # chapter_notes_status
        "E": 25,  # final_review_notes_status
        "F": 40,  # notes_on_outline_after
        "G": 40,  # final_review_notes
        "H": 15,  # sheet_row_id
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = "A3"

    # Save file
    wb.save(filename)
    logger.info(f"✓ Created Excel template: {filename}")
    logger.info(f"  - Headers in row 1 (blue)")
    logger.info(f"  - Descriptions in row 2 (gray)")
    logger.info(f"  - 3 example rows (rows 3-5)")
    logger.info(f"\nNext steps:")
    logger.info(f"  1. Open {filename} in Excel")
    logger.info(f"  2. Edit/Add your books")
    logger.info(f"  3. Run: python sync_from_excel.py")
    logger.info(f"  4. Run: python worker.py --once")


if __name__ == "__main__":
    import sys

    filename = sys.argv[1] if len(sys.argv) > 1 else "books.xlsx"
    create_excel_template(filename)
