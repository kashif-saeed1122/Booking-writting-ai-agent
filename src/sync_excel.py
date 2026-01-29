"""
Excel sync module.
Reads rows from local Excel file (.xlsx) and upserts them into Supabase books table.
Much simpler than Google Sheets - no authentication needed!
"""
import logging
import os
from typing import Dict, List, Optional

from openpyxl import load_workbook

from src.deps import get_supabase

logger = logging.getLogger(__name__)


def excel_row_to_book_dict(row_data: List, headers: List[str]) -> Optional[Dict]:
    """
    Convert an Excel row (list of values) to a book dict using header mapping.
    """
    book = {}
    for i, header in enumerate(headers):
        if i < len(row_data):
            value = row_data[i]
            # Handle None/empty cells
            if value is None:
                value = ""
            elif not isinstance(value, str):
                value = str(value)
            book[header] = value.strip() if value else ""
        else:
            book[header] = ""

    # Skip rows without title
    if not book.get("title"):
        return None

    return book


def sync_excel_to_supabase(
    excel_file: str,
    worksheet_name: str = "Books",
    header_row: int = 1,
    start_row: int = 3,  # Skip row 2 (descriptions) if template was used
) -> List[str]:
    """
    Sync Excel rows to Supabase books table.

    Expected columns (case-insensitive):
    - title (required)
    - notes_on_outline_before (required)
    - status_outline_notes (yes/no/no_notes_needed)
    - notes_on_outline_after
    - chapter_notes_status (yes/no/no_notes_needed)
    - final_review_notes_status (yes/no/no_notes_needed)
    - final_review_notes
    - sheet_row_id (optional, for tracking)

    Returns list of book_ids that were created/updated.
    """
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"Excel file not found: {excel_file}")

    logger.info(f"Loading Excel file: {excel_file}")
    wb = load_workbook(excel_file, data_only=True)
    
    # Auto-detect worksheet if not found
    if worksheet_name not in wb.sheetnames:
        available = wb.sheetnames
        logger.warning(f"Worksheet '{worksheet_name}' not found. Available: {available}")
        if available:
            worksheet_name = available[0]
            logger.info(f"Using first available worksheet: '{worksheet_name}'")
        else:
            raise ValueError("No worksheets found in Excel file")

    ws = wb[worksheet_name]

    # Read headers from first row
    headers_raw = [cell.value for cell in ws[header_row]]
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in headers_raw]

    logger.info(f"Found columns: {headers}")

    supabase = get_supabase()
    book_ids = []

    # Read data rows
    for idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not any(row):  # Skip completely empty rows
            continue

        book_dict = excel_row_to_book_dict(list(row), headers)
        if not book_dict:
            logger.warning(f"Row {idx}: Skipping - missing title")
            continue

        title = book_dict.get("title")
        if not title:
            continue

        # Skip description rows (common patterns from template)
        title_lower = str(title).lower()
        if any(keyword in title_lower for keyword in ["required", "optional", "default:", "yes/no"]):
            logger.debug(f"Row {idx}: Skipping description row")
            continue

        # Map sheet_row_id if present and valid, otherwise use row number
        sheet_row_id_raw = book_dict.get("sheet_row_id") or ""
        try:
            # Try to convert to int if it's a valid number
            if sheet_row_id_raw and str(sheet_row_id_raw).strip().isdigit():
                sheet_row_id = int(sheet_row_id_raw)
            else:
                sheet_row_id = idx
        except (ValueError, TypeError):
            sheet_row_id = idx

        # Check if book already exists (by title or sheet_row_id)
        existing = (
            supabase.table("books")
            .select("id")
            .or_(f"title.eq.{title},sheet_row_id.eq.{sheet_row_id}")
            .maybe_single()
            .execute()
        )

        book_id = None
        if existing and existing.data:
            book_id = existing.data["id"]
            logger.info(f"Row {idx}: Updating existing book '{title}' ({book_id})")
        else:
            logger.info(f"Row {idx}: Creating new book '{title}'")

        # Prepare upsert payload
        upsert_data = {
            "title": title,
            "notes_on_outline_before": book_dict.get("notes_on_outline_before", ""),
            "sheet_row_id": sheet_row_id,
        }

        # Optional fields
        for field in [
            "notes_on_outline_after",
            "status_outline_notes",
            "chapter_notes_status",
            "final_review_notes_status",
            "final_review_notes",
        ]:
            if field in book_dict and book_dict[field]:
                upsert_data[field] = book_dict[field]

        # Set defaults if not provided
        if "status_outline_notes" not in upsert_data:
            upsert_data["status_outline_notes"] = "no_notes_needed"
        if "chapter_notes_status" not in upsert_data:
            upsert_data["chapter_notes_status"] = "no_notes_needed"
        if "final_review_notes_status" not in upsert_data:
            upsert_data["final_review_notes_status"] = "no_notes_needed"
        if "outline_status" not in upsert_data:
            upsert_data["outline_status"] = "pending"
        if "book_output_status" not in upsert_data:
            upsert_data["book_output_status"] = "not_started"

        if book_id:
            supabase.table("books").update(upsert_data).eq("id", book_id).execute()
        else:
            result = supabase.table("books").insert(upsert_data).execute()
            book_id = result.data[0]["id"] if result.data else None

        if book_id:
            book_ids.append(book_id)

    logger.info(f"Sync complete. Processed {len(book_ids)} books.")
    return book_ids


if __name__ == "__main__":
    import sys

    logging.basicConfig(level=logging.INFO)
    excel_file = sys.argv[1] if len(sys.argv) > 1 else "books.xlsx"
    sync_excel_to_supabase(excel_file)
